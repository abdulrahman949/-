#importing libraries============================
import tkinter as tk
from tkinter import COMMAND, ttk
from datetime import *
from datetime import datetime ,timedelta
import calendar
from tkinter import font
import arabic_reshaper
from bidi.algorithm import get_display
import openpyxl
from openpyxl import Workbook,load_workbook
from ttkthemes import ThemedTk
from tkinter import messagebox
#opening the sheet================================

wb=Workbook()
ws=wb.active
#root=ThemedTk(theme='scidblue')
root=tk.Tk()

root.title('جدول تسجيل الحضور بمركز الإمام')
tableframe=ttk.Frame(root,height=300,width=300)
#headings==========================================
month = datetime.now().month
year = datetime.now().year
today = datetime.today()
wtoday=today.strftime('%A')
tabledays=[]
# حساب بداية الأسبوع الحالي (نفترض أن الأسبوع يبدأ من يوم السبت)
start_of_week = today - timedelta(days=(today.weekday() + 2) % 7)
satdate=start_of_week.day
end_of_week= satdate + 5



# طباعة تواريخ الأيام في الأسبوع الحالي
weekdays=[]
for i in range(5):
	weekdays.append(satdate+i)
	if i==0:
		weekdays[i]='السبت'+str(weekdays[i])
	elif i==1:	
		weekdays[i]='الأحد'+str(weekdays[i])
	elif i==2:
		weekdays[i]='الإثنين'+str(weekdays[i])
	elif i==3:
		weekdays[i]='الثلاثاء'+str(weekdays[i])
	elif i==4:
		weekdays[i]='الأربعاء'+str(weekdays[i])



empty_headings=['الكود','الإسم','قيمة الإشتراك','السداد']

headings=[]

for i in weekdays:
    empty_headings.append(i)
for i in range (1,10):
	headings.append(empty_headings[-i])

treescroll=ttk.Scrollbar(tableframe)
tree = ttk.Treeview(tableframe , columns=headings,yscrollcommand=treescroll.set, show='headings',height=14,selectmode='browse')

style= ttk.Style()
style.configure('Treeview',rowheight=25, borderwidth=1 ,relief='solid')
style.map('Treeview',background=[('selected','#006400')],foreground=[('selected','white')])

#table window=====================================
def load():
	treescroll.pack(side='right',fill='y')
	for i in headings:
		tree.heading(i, text=i)
		tree.column(i, width=50,anchor='center')
		tree.column('#8', width=150,anchor='e')
		tree.column('#7', width=70,anchor='center')

	workbook = openpyxl.load_workbook("E:\data.xlsx.xlsx")
	sheet=workbook.active
	list_values=list(sheet.values)
	for j in list_values:
		final=[]
		for i in j:
			if i == None :
				final.append('')
			else:
				final.append(i)
		tree.insert("","end",values=final)
	tree.pack(side='left',fill='both',expand='True')
	workbook.save("E:\data.xlsx.xlsx")







def reload():
	for item in tree.get_children():
		tree.delete(item)
	load()
#add function =====================================
def add():
	global tree
	global ws
	global wb
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	values=[]
	code=codeentry2.get()
	name=nameentry.get()
	for i in (worksheet['I']):
		values.append(str(i.value))
	if code not in values:
		try:
			int(code)
			worksheet.append([None,None,None,None,None,None,None,name,int(code)])
			tree.insert("","end",values=['','','','','','','',name,code])
			workbook.save("E:\data.xlsx.xlsx")
		except:
			messagebox.showwarning('تحذير',message='كود غير صالح')
			codeentry2.delete(0,"end")
			nameentry.delete(0,"end")
			codeentry2.insert('0','الكود')
			nameentry.insert('0','الإسم الجديد')
	else:
		messagebox.showwarning('تحذير',message='كود موجود بالفعل')







#attendance===================================





def record(event=None):
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	workbook2=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook2.active
	coderecord=codeentry.get()
	try :
		if wtoday == 'Saturday':
			column2='E'
		elif wtoday == 'Sunday':
			column2='D'
		elif wtoday == 'Monday':
			column2='C'
		elif wtoday == 'Tuesday':
			column2='B'
		elif wtoday == 'Wednesday':
			column2='A'
		int(coderecord)
		values=[]
		for v in worksheet["I"]:
			values.append(v.value)
		for i in values:
			if int(i) == int(coderecord):
				row2=(values).index(i)
			else :
				pass
		worksheet[f'{column2}{row2+1}']='حضور'
		workbook2.save("E:\data.xlsx.xlsx")
		codeentry.delete(0,"end")
		reload()
	except:
		messagebox.showwarning('تحذير',message='كود غير صالح')
		codeentry.delete('0','end')
		codeentry.insert('0','الكود')


#edit=========================================

def edit():
	global wb
	global ws
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	codeedit=codeentry2.get()
	nameedit=nameentry.get()
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	try:
		values=[]
		for v in worksheet["I"]:
			values.append(v.value)
			for i in values:
				if int(i) == int(codeedit):
					row2=(values).index(i)
				else :
					pass
		worksheet.cell(column=8,row=row2+1,value=nameedit)
		workbook.save("E:\data.xlsx.xlsx")
		reload()
	except:
		messagebox.showwarning('تحذير',message='كود غير صالح')
		codeentry2.delete('0','end')
		nameentry.delete('0','end')
		codeentry2.insert('0','الكود')
		nameentry.insert('0','الإسم الجديد')
#=============
#money========================================

def money(event=None):
	global wb
	global ws
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	codemoney=codeentry3.get()
	try:
		values=[]
		for v in worksheet["I"]:
			values.append(v.value)
			for i in values:
				if int(i) == int(codemoney):
					row2=(values).index(i)
				else :
					pass
		worksheet.cell(column=6,row=row2+1,value='تم الدفع')
		workbook.save("E:\data.xlsx.xlsx")
		reload()
	except:
		messagebox.showwarning('تحذير',message='كود غير صالح')
		codeentry3.delete('0','end')
		codeentry3.insert('0','الكود')

# check===========================================================

def check(entry,string):
	if len(entry.get())==0:
		entry.insert('0',string)
	else:
		pass

def clear(event=None):
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	password=passwordentry.get()
	if password=='0000':
		for row in range(1,worksheet.max_row +1 ):
			worksheet[f'F{row}'].value=None
		workbook.save("E:\data.xlsx.xlsx")
		passwordentry.delete('0','end')
		messagebox.showinfo('تنبيه',message='تم التفريغ بنجاح')
		reload()
	else:
		messagebox.showerror('تحذير',message='كلمة سر خاطئة')
#money amount ====================================================

def moneyamount(event=None):
	workbook=openpyxl.load_workbook('E:\data.xlsx.xlsx')
	worksheet=workbook.active
	codemoney=codeentry3.get()
	moneyamounty=moneyamountentry.get()
	try:
		values=[]
		for v in worksheet["I"]:
			values.append(v.value)
			for i in values:
				if int(i) == int(codemoney):
					row2=(values).index(i)
				else :
					pass
		worksheet.cell(column=7,row=row2+1,value=moneyamounty)
		workbook.save("E:\data.xlsx.xlsx")
		reload()
	except:
		messagebox.showwarning('تحذير',message='رقم غير صالح')
		codeentry3.delete('0','end')
		moneyamountentry.delete('0','end')
		codeentry3.insert('0','الكود')
		moneyamountentry.insert('0','المبلغ')

#=============
#التنسيق===================================================================
#frame==============================================
entryframe=ttk.Frame(root)
frametitle=get_display(arabic_reshaper.reshape('أدخل المعلومات'))
labelframe=ttk.LabelFrame(entryframe,text=frametitle)
title1=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('تسجيل الحضور ')))
title3=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('تسجيل الإشتراك ')))
title2=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('الإضافه والتعديل')))
title2.grid(row=9,column=0,padx=30)
code1label=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('الكود')))
code2label=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('الكود')))
namelabel=ttk.Label(labelframe, text=get_display(arabic_reshaper.reshape('الاسم الجديد')))
title3.grid(row=4,column=0,padx=30)
title1.grid(row=0,column=0,padx=25,pady=5)
entryframe.grid(row=0,column=0)
codeentry=ttk.Entry(labelframe,width=20)
codeentry2=ttk.Entry(labelframe)
codeentry.grid(row=1,column=0,padx=5,pady=5,sticky='w')
sep1=ttk.Separator(labelframe)
sep2=ttk.Separator(labelframe)
sep1.grid(row=3,column=0,sticky='we',padx=10,pady=1)
sep3=ttk.Separator(labelframe)
sep3.grid(row=14,column=0,sticky='we',padx=10,pady=10)
sep2.grid(row=8,column=0,sticky='we',padx=10,pady=1)
nameentry=ttk.Entry(labelframe)
nameentry.grid(row=11,column=0,padx=5,pady=1,sticky='w')
codeentry2.grid(row=10,column=0,padx=5,pady=1,sticky='w')
codeentry3=ttk.Entry(labelframe)
codeentry3.grid(row=5,column=0,padx=5,pady=1,sticky='w')
addbutton = ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('اضافه')),command=add)
addbutton.grid(row=12,column=0,padx=10,pady=1,sticky='we')
editbutton = ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('تعديل')),command=edit)
editbutton.grid(row=13,column=0,padx=10,pady=1,sticky='we')
attendancebutton = ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('تسجيل')),command=record)
moneybutton = ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('دفع')),command=money)
moneyamountbutton = ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('إضافه')),command=moneyamount)
moneyamountentry=ttk.Entry(labelframe,width=8)
moneyamountentry.grid(row=7,column=0,padx=5,pady=1,sticky='w')
moneyamountbutton.grid(row=7,column=0,padx=5,pady=5,sticky='e')
moneybutton.grid(row=6,column=0,padx=5,pady=1,sticky='we')
attendancebutton.grid(row=2,column=0,padx=5,pady=1,sticky='we')
labelframe.grid(row=1,column=0,padx=10,pady=10)
tableframe.grid(row=0,column=1)
clearbutton=ttk.Button(labelframe,text=get_display(arabic_reshaper.reshape('تفريغ')),command=clear)
passwordentry=ttk.Entry(labelframe,width=10)
clearbutton.grid(row=15,column=0,sticky='e')
passwordentry.grid(row=15,column=0,sticky='w')
treescroll.config(command=tree.yview)
codeentry.bind ('<Return>', record )
moneyamountentry.bind ('<Return>', moneyamount )
codeentry3.bind ('<Return>', money )
passwordentry.bind ('<Return>', clear )



codeentry.bind('<FocusIn>',lambda e:codeentry.delete('0','end'))
codeentry2.bind('<FocusIn>',lambda e:codeentry2.delete('0','end'))
codeentry3.bind('<FocusIn>',lambda e:codeentry3.delete('0','end'))
nameentry.bind('<FocusIn>',lambda e:nameentry.delete('0','end'))
moneyamountentry.bind('<FocusIn>',lambda e:moneyamountentry.delete('0','end'))
codeentry.bind('<FocusOut>',lambda e:check(codeentry,'الكود'))
codeentry2.bind('<FocusOut>',lambda e:check(codeentry2,'الكود'))
codeentry3.bind('<FocusOut>',lambda e:check(codeentry3,'الكود'))
nameentry.bind('<FocusOut>',lambda e:check(nameentry,'الإسم الجديد'))
passwordentry.bind('<FocusOut>',lambda e:check(passwordentry,'كلمة السر'))
passwordentry.bind('<FocusIn>',lambda e:passwordentry.delete('0','end'))

codeentry.insert('0','الكود')
codeentry2.insert('0','الكود')
codeentry3.insert('0','الكود')
nameentry.insert('0','الإسم الجدبد')
moneyamountentry.insert('0','المبلغ')
passwordentry.insert('0','كلمة السر')

load()



#main window========================================

root.mainloop()
